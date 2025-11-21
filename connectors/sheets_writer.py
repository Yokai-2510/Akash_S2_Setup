"""
Data Writer Connector - Professional Template Generator.
Creates beautiful client-facing Excel with hidden _SYSTEM_DATA sheet for Python.
"""
import pandas as pd
import os
from datetime import datetime
from typing import Dict, Any
from utils.logger import setup_logger

log = setup_logger()

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False


def write_all_sheets_to_excel(universal_data: Dict[str, Any]) -> Dict[str, Any]:
    """Main entry point."""
    mode = universal_data['configs']['system_settings']['system'].get('data_source_mode', 'local_excel')
    log.info(f"=== WRITING DATA ({mode.upper()}) ===", tags=["WRITER", "START"])
    
    try:
        if mode == 'local_excel':
            _write_to_local_excel(universal_data)
        log.info("=== WRITING COMPLETE ===", tags=["WRITER", "SUCCESS"])
    except Exception as e:
        log.critical(f"Write failed: {e}", tags=["WRITER", "ERROR"], exc_info=True)
        raise
    return universal_data


def _write_to_local_excel(universal_data: Dict[str, Any]):
    path_config = universal_data['configs']['system_settings']['paths']
    file_path = os.path.join(universal_data['system']['project_root'], path_config.get('local_excel_file', ''))
    
    if not os.path.exists(file_path):
        if universal_data['configs']['system_settings']['system'].get('create_missing_local_excel', False):
            log.warning("Creating professional Excel template...", tags=["WRITER", "TEMPLATE"])
            _create_professional_template(file_path, universal_data)
            return
        else:
            raise FileNotFoundError(f"Excel file missing: {file_path}")
    
    _update_existing_workbook(file_path, universal_data)


def _create_professional_template(file_path: str, universal_data: Dict[str, Any] = None):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    if not HAS_XLSXWRITER:
        log.error("xlsxwriter not installed. Install with: pip install xlsxwriter", tags=["WRITER"])
        raise ImportError("xlsxwriter required for template creation")
    
    # Get defaults from strategy_config if available
    default_etfs = [
        ('ETF_01', 'NIFTYBEES', True, 30.0, '', 'Core,Broad'),
        ('ETF_02', 'BANKBEES', True, 15.0, 2.5, 'Core,Financials'),
        ('ETF_03', 'GOLDBEES', True, 12.0, 3.0, 'Satellite,Commodity'),
        ('ETF_04', 'JUNIORBEES', True, 10.0, '', 'Satellite,Midcap'),
        ('ETF_05', 'ITBEES', True, 8.0, 2.8, 'Satellite,Sector'),
        ('ETF_06', 'SILVERBEES', True, 8.0, 3.5, 'Satellite,Commodity'),
        ('ETF_07', 'CPSEETF', True, 7.0, '', 'Satellite,PSU'),
        ('ETF_08', 'PHARMABEES', True, 5.0, 3.0, 'Satellite,Sector'),
        ('ETF_09', 'MOM100', True, 5.0, 2.5, 'Satellite,Momentum'),
    ]
    
    alloc_defaults = {'S2_Target_%': 34.0, 'Weeks_to_Glide': 52, 'Weekly_Transfer_Cap_%': 5.0,
                      'S2_Weekly_Budget_Cap_%': 1.25, 'Enable_Carry_Forward': 'TRUE',
                      'Initial_Capital': 1000000}  # Added initial capital
    risk_defaults = {'DriftBand_%': 10.0, 'CoreFloor_%': 70.0, 'ATR_Ceiling_%': 2.0,
                     'Single_ETF_Max_%': 35.0, 'MaxTrimPerETF_%': 25.0, 'WeeklyHarvestCap_%': 12.0}
    
    # Override from strategy_config if available
    if universal_data:
        strat = universal_data['configs'].get('strategy_settings', {})
        alloc = strat.get('allocation_rules', {})
        risk = strat.get('risk_controls', {})
        
        if alloc.get('s2_target_percent'): alloc_defaults['S2_Target_%'] = alloc['s2_target_percent']
        if alloc.get('weeks_to_glide'): alloc_defaults['Weeks_to_Glide'] = alloc['weeks_to_glide']
        if alloc.get('weekly_transfer_cap_percent'): alloc_defaults['Weekly_Transfer_Cap_%'] = alloc['weekly_transfer_cap_percent']
        if alloc.get('s2_weekly_budget_cap_percent'): alloc_defaults['S2_Weekly_Budget_Cap_%'] = alloc['s2_weekly_budget_cap_percent']
        
        if risk.get('drift_band_percent'): risk_defaults['DriftBand_%'] = risk['drift_band_percent']
        if risk.get('core_floor_percent'): risk_defaults['CoreFloor_%'] = risk['core_floor_percent']
        if risk.get('default_atr_ceiling_percent'): risk_defaults['ATR_Ceiling_%'] = risk['default_atr_ceiling_percent']
        if risk.get('single_etf_max_percent'): risk_defaults['Single_ETF_Max_%'] = risk['single_etf_max_percent']
        if risk.get('max_trim_per_etf_percent'): risk_defaults['MaxTrimPerETF_%'] = risk['max_trim_per_etf_percent']
        if risk.get('weekly_harvest_cap_percent'): risk_defaults['WeeklyHarvestCap_%'] = risk['weekly_harvest_cap_percent']
        
        # ETFs from strategy config
        etf_list = strat.get('universe', {}).get('etfs_to_track', [])
        if etf_list:
            default_etfs = [(f'ETF_{str(i+1).zfill(2)}', etf, True, round(100/len(etf_list), 1), '', 'Core') 
                           for i, etf in enumerate(etf_list)]
    
    wb = xlsxwriter.Workbook(file_path)
    
    # === STYLES ===
    fmt = {
        'title': wb.add_format({'bold': True, 'font_size': 18, 'font_color': '#1F4E79', 'bottom': 2}),
        'subtitle': wb.add_format({'bold': True, 'font_size': 11, 'font_color': '#5B9BD5', 'italic': True}),
        'section': wb.add_format({'bold': True, 'font_size': 11, 'bg_color': '#1F4E79', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'}),
        'header': wb.add_format({'bold': True, 'bg_color': '#D6DCE4', 'border': 1, 'align': 'center', 'text_wrap': True}),
        'data': wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'}),
        'data_left': wb.add_format({'border': 1, 'align': 'left'}),
        'currency': wb.add_format({'border': 1, 'align': 'right', 'num_format': '₹#,##0.00'}),
        'pct': wb.add_format({'border': 1, 'align': 'center', 'num_format': '0.00'}),
        'input': wb.add_format({'border': 2, 'bg_color': '#FFFFCC', 'align': 'center'}),
        'locked': wb.add_format({'border': 1, 'bg_color': '#F2F2F2', 'align': 'center'}),
        'good': wb.add_format({'border': 1, 'align': 'center', 'bg_color': '#C6EFCE', 'font_color': '#006100'}),
        'bad': wb.add_format({'border': 1, 'align': 'center', 'bg_color': '#FFC7CE', 'font_color': '#9C0006'}),
        'warn': wb.add_format({'border': 1, 'align': 'center', 'bg_color': '#FFEB9C', 'font_color': '#9C5700'}),
        'kpi_label': wb.add_format({'bold': True, 'font_size': 10, 'align': 'right'}),
        'kpi_val': wb.add_format({'bold': True, 'font_size': 14, 'font_color': '#1F4E79'}),
    }
    
    default_etfs = [
        ('ETF_01', 'NIFTYBEES', True, 40.0, '', 'Core,Broad'),
        ('ETF_02', 'BANKBEES', True, 20.0, 2.5, 'Core,Financials'),
        ('ETF_03', 'GOLDBEES', True, 15.0, 3.0, 'Satellite,Commodity'),
        ('ETF_04', 'JUNIORBEES', True, 10.0, '', 'Satellite,Midcap'),
        ('ETF_05', 'ITBEES', True, 6.0, 2.8, 'Satellite,Sector'),
        ('ETF_06', 'SILVERBEES', True, 6.0, 3.5, 'Satellite,Commodity'),
        ('ETF_07', 'MON100', False, 3.0, '', 'Satellite,Momentum'),
    ]
    
    # ========== 1. DASHBOARD ==========
    ws = wb.add_worksheet('DASHBOARD')
    ws.set_column('A:A', 2)
    ws.set_column('B:J', 14)
    
    ws.merge_range('B2:J2', 'S2 TRADING SYSTEM - WEEKLY DASHBOARD', fmt['title'])
    ws.write('B3', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', fmt['subtitle'])
    
    # KPI Row 1
    for col, (label, cell) in enumerate([('TOTAL CAPITAL', 'B'), ('S2 CURRENT %', 'D'), ('S2 TARGET %', 'F'), ('GAP TO TARGET', 'H')]):
        ws.merge_range(f'{cell}5:{chr(ord(cell)+1)}5', label, fmt['section'])
    ws.merge_range('B6:C6', '₹ 0', fmt['kpi_val'])
    ws.merge_range('D6:E6', '0.00%', fmt['kpi_val'])
    ws.merge_range('F6:G6', '34.00%', fmt['kpi_val'])
    ws.merge_range('H6:I6', '₹ 0', fmt['kpi_val'])
    
    # KPI Row 2
    for col, (label, cell) in enumerate([('WEEKLY BUDGET', 'B'), ('ACCRUED CARRY', 'D'), ('ETFs HEALTHY', 'F'), ('ACTIONS TODAY', 'H')]):
        ws.merge_range(f'{cell}8:{chr(ord(cell)+1)}8', label, fmt['section'])
    ws.merge_range('B9:C9', '₹ 0', fmt['kpi_val'])
    ws.merge_range('D9:E9', '₹ 0', fmt['kpi_val'])
    ws.merge_range('F9:G9', '0 / 7', fmt['kpi_val'])
    ws.merge_range('H9:I9', '0 BUY | 0 TRIM', fmt['kpi_val'])
    
    # Portfolio Table
    ws.merge_range('B11:J11', 'PORTFOLIO OVERVIEW', fmt['section'])
    headers = ['Ticker', 'Current %', 'Target %', 'Gap %', 'Status', 'Health', 'Signal', 'Price', 'Action']
    ws.write_row('B12', headers, fmt['header'])
    for r in range(13, 20):
        for c in range(1, 10):
            ws.write(r, c, '', fmt['data'])
    
    # ========== 2. CONFIG ==========
    ws = wb.add_worksheet('CONFIG')
    ws.set_column('A:A', 2)
    ws.set_column('B:B', 26)
    ws.set_column('C:C', 14)
    ws.set_column('D:D', 8)
    ws.set_column('E:E', 45)
    ws.set_column('F:G', 14)
    
    ws.merge_range('B2:E2', 'S2 STRATEGY CONFIGURATION', fmt['title'])
    ws.write('B3', '⚠ Edit YELLOW cells only. Changes auto-sync to system.', fmt['subtitle'])
    
    # Section A
    ws.merge_range('B5:E5', 'ALLOCATION RULES', fmt['section'])
    ws.write_row('B6', ['Parameter', 'Value', 'Unit', 'Description'], fmt['header'])
    alloc = [
        ('Initial_Capital', alloc_defaults['Initial_Capital'], '₹', 'Total portfolio capital available'),
        ('S2_Target_%', alloc_defaults['S2_Target_%'], '%', 'Target S2 allocation from Asset Allocator'),
        ('Weeks_to_Glide', alloc_defaults['Weeks_to_Glide'], 'weeks', 'Weeks to close gap to target'),
        ('Weekly_Transfer_Cap_%', alloc_defaults['Weekly_Transfer_Cap_%'], '%', 'Max % of portfolio to transfer weekly'),
        ('S2_Weekly_Budget_Cap_%', alloc_defaults['S2_Weekly_Budget_Cap_%'], '%', 'Max % of S2 to deploy weekly'),
        ('Enable_Carry_Forward', alloc_defaults['Enable_Carry_Forward'], 'bool', 'Roll unused budget to next week'),
    ]
    for i, (p, v, u, d) in enumerate(alloc):
        ws.write(6+i, 1, p, fmt['data_left'])
        ws.write(6+i, 2, v, fmt['input'])
        ws.write(6+i, 3, u, fmt['data'])
        ws.write(6+i, 4, d, fmt['data_left'])
    
    # Section B
    ws.merge_range('B13:E13', 'RISK CONTROLS', fmt['section'])
    ws.write_row('B14', ['Parameter', 'Value', 'Unit', 'Description'], fmt['header'])
    risk = [
        ('DriftBand_%', risk_defaults['DriftBand_%'], '%', 'Tolerance before rebalancing action'),
        ('CoreFloor_%', risk_defaults['CoreFloor_%'], '%', 'Min position size as % of target'),
        ('ATR_Ceiling_%', risk_defaults['ATR_Ceiling_%'], '%', 'Max acceptable volatility'),
        ('Single_ETF_Max_%', risk_defaults['Single_ETF_Max_%'], '%', 'Max single ETF concentration'),
        ('MaxTrimPerETF_%', risk_defaults['MaxTrimPerETF_%'], '%', 'Max trim per harvest event'),
        ('WeeklyHarvestCap_%', risk_defaults['WeeklyHarvestCap_%'], '%', 'Weekly profit-taking limit'),
    ]
    for i, (p, v, u, d) in enumerate(risk):
        ws.write(14+i, 1, p, fmt['data_left'])
        ws.write(14+i, 2, v, fmt['input'])
        ws.write(14+i, 3, u, fmt['data'])
        ws.write(14+i, 4, d, fmt['data_left'])
    
    # Section C
    ws.merge_range('B22:G22', 'ETF LINEUP (Target % must sum to 100)', fmt['section'])
    ws.write_row('B23', ['ETF_ID', 'Ticker', 'Enabled', 'Target_%', 'ATR_Override', 'Tags'], fmt['header'])
    for i, (eid, ticker, en, tgt, atr, tags) in enumerate(default_etfs):
        r = 23 + i
        ws.write(r, 1, eid, fmt['locked'])
        ws.write(r, 2, ticker, fmt['input'])
        ws.write(r, 3, 'TRUE' if en else 'FALSE', fmt['good'] if en else fmt['bad'])
        ws.write(r, 4, tgt, fmt['input'])
        ws.write(r, 5, atr if atr else '', fmt['input'])
        ws.write(r, 6, tags, fmt['input'])
    
    # ========== 3. STATE ==========
    ws = wb.add_worksheet('STATE')
    ws.set_column('A:A', 2)
    ws.set_column('B:L', 13)
    
    ws.merge_range('B2:K2', 'PORTFOLIO STATE', fmt['title'])
    headers = ['ETF_ID', 'Ticker', 'Units', 'Avg_Cost', 'Current_Price', 'Market_Value', 'Current_%', 'Target_%', 'Gap_%', 'Status']
    ws.write_row('B4', headers, fmt['header'])
    for r in range(5, 12):
        for c in range(1, 11):
            ws.write(r, c, '', fmt['data'])
    
    ws.merge_range('B14:C14', 'SUMMARY', fmt['section'])
    ws.write('B15', 'Total S2 Value:', fmt['kpi_label'])
    ws.write('C15', '₹ 0', fmt['kpi_val'])
    ws.write('B16', 'Current Weight:', fmt['kpi_label'])
    ws.write('C16', '0%', fmt['kpi_val'])
    
    # ========== 4. SIGNALS ==========
    ws = wb.add_worksheet('SIGNALS')
    ws.set_column('A:A', 2)
    ws.set_column('B:M', 11)
    
    ws.merge_range('B2:L2', 'TECHNICAL SIGNALS & HEALTH GATES', fmt['title'])
    headers = ['Week', 'ETF_ID', 'Ticker', 'TSI', 'TSI_Sig', 'RSI', 'VWMA_Slope', 'ATR%', 'ATR_Cap', 'Health', 'Pass']
    ws.write_row('B4', headers, fmt['header'])
    
    ws.merge_range('B18:F18', 'HEALTH GATE DEFINITIONS', fmt['section'])
    gates = [('G1: TSI > Signal', 'Trend up'), ('G2: RSI > 50', 'Momentum +'), ('G3: VWMA↑', 'Vol trend'), ('G4: ATR ≤ Cap', 'Low vol')]
    for i, (g, d) in enumerate(gates):
        ws.write(18+i, 1, g, fmt['data_left'])
        ws.write(18+i, 2, d, fmt['data_left'])
    
    # ========== 5. WEEKLY_ACTIONS ==========
    ws = wb.add_worksheet('WEEKLY_ACTIONS')
    ws.set_column('A:A', 2)
    ws.set_column('B:L', 12)
    
    ws.merge_range('B2:K2', 'WEEKLY ACTION LOG', fmt['title'])
    headers = ['Week', 'ETF_ID', 'Ticker', 'Action', 'Units', 'Price', 'Amount', 'Reason', 'Pre_%', 'Post_%']
    ws.write_row('B4', headers, fmt['header'])
    
    # ========== 6. HARVEST_LOG ==========
    ws = wb.add_worksheet('HARVEST_LOG')
    ws.set_column('A:A', 2)
    ws.set_column('B:H', 14)
    
    ws.merge_range('B2:G2', 'HARVEST (TRIM) LOG', fmt['title'])
    headers = ['Date', 'Ticker', 'Trigger', 'Units', 'Price', 'Amount']
    ws.write_row('B4', headers, fmt['header'])
    
    # ========== 7. SYSTEM_CONTROL ==========
    ws = wb.add_worksheet('SYSTEM_CONTROL')
    ws.set_column('A:A', 2)
    ws.set_column('B:D', 22)
    
    ws.merge_range('B2:D2', 'SYSTEM CONTROL PANEL', fmt['title'])
    ws.write('B3', 'Set UPDATE_TRIGGER=TRUE to run pipeline', fmt['subtitle'])
    
    ws.write_row('B5', ['Parameter', 'Value', 'Updated'], fmt['header'])
    ctrl = [('UPDATE_TRIGGER', 'FALSE'), ('FORCE_FULL_REFRESH', 'FALSE'), ('RUN_STATUS', 'IDLE'), ('LAST_RUN_DATE', ''), ('ERROR_MESSAGE', '')]
    for i, (p, v) in enumerate(ctrl):
        ws.write(5+i, 1, p, fmt['data_left'])
        ws.write(5+i, 2, v, fmt['input'] if 'TRIGGER' in p else fmt['data'])
        ws.write(5+i, 3, '', fmt['data'])
    
    # ========== 8. _SYSTEM_DATA (HIDDEN) ==========
    ws = wb.add_worksheet('_SYSTEM_DATA')
    ws.hide()
    
    ws.write_row(0, 0, ['KEY', 'VALUE'], fmt['header'])
    params = [
        ('Initial_Capital', alloc_defaults['Initial_Capital']),
        ('S2_Target_%', alloc_defaults['S2_Target_%']),
        ('Weeks_to_Glide', alloc_defaults['Weeks_to_Glide']),
        ('Weekly_Transfer_Cap_%', alloc_defaults['Weekly_Transfer_Cap_%']),
        ('S2_Weekly_Budget_Cap_%', alloc_defaults['S2_Weekly_Budget_Cap_%']),
        ('Enable_Carry_Forward', alloc_defaults['Enable_Carry_Forward']),
        ('DriftBand_%', risk_defaults['DriftBand_%']),
        ('CoreFloor_%', risk_defaults['CoreFloor_%']),
        ('ATR_Ceiling_%', risk_defaults['ATR_Ceiling_%']),
        ('Single_ETF_Max_%', risk_defaults['Single_ETF_Max_%']),
        ('MaxTrimPerETF_%', risk_defaults['MaxTrimPerETF_%']),
        ('WeeklyHarvestCap_%', risk_defaults['WeeklyHarvestCap_%']),
    ]
    for i, (k, v) in enumerate(params):
        ws.write(1+i, 0, k)
        ws.write(1+i, 1, v)
    
    # ETF Lineup
    lr = len(params) + 3
    ws.write(lr, 0, '### ETF_LINEUP ###')
    ws.write_row(lr+1, 0, ['ETF_ID', 'Ticker', 'Enabled', 'Target_%', 'ATR_Override_%', 'Tags'])
    for i, etf in enumerate(default_etfs):
        ws.write_row(lr+2+i, 0, [etf[0], etf[1], 'TRUE' if etf[2] else 'FALSE', etf[3], etf[4] if etf[4] else '', etf[5]])
    
    # Control
    cr = lr + 2 + len(default_etfs) + 2
    ws.write(cr, 0, '### CONTROL ###')
    ws.write_row(cr+1, 0, ['UPDATE_TRIGGER', 'FALSE'])
    ws.write_row(cr+2, 0, ['RUN_STATUS', 'IDLE'])
    ws.write_row(cr+3, 0, ['LAST_RUN_DATE', ''])
    ws.write_row(cr+4, 0, ['ERROR_MESSAGE', ''])
    
    wb.close()
    log.info(f"✓ Professional template created: {file_path}", tags=["WRITER", "SUCCESS"])


def _update_existing_workbook(file_path: str, universal_data: Dict[str, Any]):
    """Updates workbook with pipeline results."""
    from openpyxl import load_workbook
    
    report = universal_data.get('report_sheets', {})
    analysis = universal_data.get('analysis', {})
    portfolio = universal_data.get('portfolio_state', {})
    summary = portfolio.get('summary', {})
    
    wb = load_workbook(file_path)
    
    # Update DASHBOARD
    if 'DASHBOARD' in wb.sheetnames:
        ws = wb['DASHBOARD']
        ws['B6'] = f"₹ {summary.get('total_s2_value', 0):,.0f}"
        ws['D6'] = f"{summary.get('current_s2_weight_pct', 0):.2f}%"
        ws['H6'] = f"₹ {analysis.get('gap_to_target', 0):,.0f}"
        ws['B9'] = f"₹ {analysis.get('weekly_budget', 0):,.0f}"
        ws['D9'] = f"₹ {analysis.get('accrued_carry', 0):,.0f}"
        
        health_df = analysis.get('health_matrix_df')
        if health_df is not None and not health_df.empty:
            passing = len(health_df[health_df['Pass'] == True])
            ws['F9'] = f"{passing} / {len(health_df)}"
        
        actions_df = report.get('weekly_actions')
        if actions_df is not None and not actions_df.empty:
            buys = len(actions_df[actions_df['Action'] == 'BUY'])
            trims = len(actions_df[actions_df['Action'] == 'TRIM'])
            ws['H9'] = f"{buys} BUY | {trims} TRIM"
        
        dash_df = report.get('dashboard')
        if dash_df is not None and not dash_df.empty:
            _write_df_to_sheet(ws, dash_df, start_row=13, start_col=2)
    
    # Update STATE
    if 'STATE' in wb.sheetnames and report.get('portfolio_state') is not None:
        ws = wb['STATE']
        _write_df_to_sheet(ws, report['portfolio_state'], start_row=5, start_col=2)
        ws['C15'] = f"₹ {summary.get('total_s2_value', 0):,.0f}"
        ws['C16'] = f"{summary.get('current_s2_weight_pct', 0):.1f}%"
    
    # Update SIGNALS
    if 'SIGNALS' in wb.sheetnames and report.get('signals') is not None:
        _write_df_to_sheet(wb['SIGNALS'], report['signals'], start_row=5, start_col=2)
    
    # Update WEEKLY_ACTIONS
    if 'WEEKLY_ACTIONS' in wb.sheetnames and report.get('weekly_actions') is not None:
        _write_df_to_sheet(wb['WEEKLY_ACTIONS'], report['weekly_actions'], start_row=5, start_col=2)
    
    # Update _SYSTEM_DATA control
    if '_SYSTEM_DATA' in wb.sheetnames:
        ws = wb['_SYSTEM_DATA']
        for row in ws.iter_rows():
            if row[0].value == 'UPDATE_TRIGGER':
                row[1].value = 'FALSE'
            elif row[0].value == 'RUN_STATUS':
                row[1].value = 'SUCCESS'
            elif row[0].value == 'LAST_RUN_DATE':
                row[1].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Update SYSTEM_CONTROL
    if 'SYSTEM_CONTROL' in wb.sheetnames:
        ws = wb['SYSTEM_CONTROL']
        ws['C6'] = 'FALSE'
        ws['C8'] = 'SUCCESS'
        ws['C9'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    wb.save(file_path)
    log.info("✓ Workbook updated", tags=["WRITER", "SUCCESS"])


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int, start_col: int):
    """Writes DataFrame to worksheet starting at given position."""
    if df is None or df.empty:
        return
    for i, (_, row) in enumerate(df.iterrows()):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j)
            cell.value = val if pd.notna(val) else ''


def update_control_cells(universal_data: Dict[str, Any], updates: dict) -> None:
    """Updates control cells in Excel."""
    from openpyxl import load_workbook
    
    path_config = universal_data['configs']['system_settings']['paths']
    file_path = os.path.join(universal_data['system']['project_root'], path_config.get('local_excel_file', ''))
    
    if not os.path.exists(file_path):
        return
    
    wb = load_workbook(file_path)
    
    # Update _SYSTEM_DATA
    if '_SYSTEM_DATA' in wb.sheetnames:
        ws = wb['_SYSTEM_DATA']
        for row in ws.iter_rows():
            if row[0].value in updates:
                row[1].value = updates[row[0].value]
    
    # Update SYSTEM_CONTROL
    if 'SYSTEM_CONTROL' in wb.sheetnames:
        ws = wb['SYSTEM_CONTROL']
        cell_map = {'UPDATE_TRIGGER': 'C6', 'RUN_STATUS': 'C8', 'LAST_RUN_DATE': 'C9', 'ERROR_MESSAGE': 'C10'}
        for key, cell in cell_map.items():
            if key in updates:
                ws[cell] = updates[key]
    
    wb.save(file_path)